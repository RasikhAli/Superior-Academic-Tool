from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from datetime import datetime
import pandas as pd
import re
import os

def convert_to_24hour(time_str, period):
    """Convert 12-hour time to 24-hour format"""
    if not period:
        return time_str

    hour, minute = time_str.split(':')
    hour = int(hour)

    if period.upper() == 'PM' and hour != 12:
        hour += 12
    elif period.upper() == 'AM' and hour == 12:
        hour = 0

    return f"{hour:02d}:{minute}"

def is_time_cell(cell_value):
    """Check if a cell contains a time value (e.g., '8:00-9:20', '10:45 AM - 12:25 PM')"""
    if not cell_value:
        return False
    cell_str = str(cell_value).strip()
    # Match patterns like "8:00-9:20", "8:00 - 9:20", "10:45 AM - 12:25 PM"
    time_pattern = r'\d{1,2}:\d{2}\s*[-–]\s*\d{1,2}:\d{2}'
    return bool(re.search(time_pattern, cell_str))

def is_day_name(cell_value):
    """Check if a cell contains a day name"""
    if not cell_value:
        return False
    cell_str = str(cell_value).strip().lower()
    days = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday']
    return cell_str in days or isinstance(cell_value, datetime)

def detect_day_blocks(ws):
    """
    Automatically detect day blocks in a worksheet.
    Returns a list of dictionaries with day_cell, time_row, room_rows, and time_cols.
    """
    blocks = []
    merged_cells = ws.merged_cells.ranges
    processed_rows = set()  # Track which rows we've already processed as day blocks

    def get_cell_value(row, col):
        """Get cell value, handling merged cells"""
        cell = ws.cell(row=row, column=col)
        coord = f"{get_column_letter(col)}{row}"
        for merged in merged_cells:
            if coord in merged:
                return ws.cell(merged.min_row, merged.min_col).value
        return cell.value

    # Search for day cells in column B (column 2)
    max_row = ws.max_row
    for row in range(1, max_row + 1):
        if row in processed_rows:
            continue

        cell_value = get_cell_value(row, 2)  # Column B

        if is_day_name(cell_value):
            # Found a potential day cell
            day_row = row

            # Check if there's a time cell in the same row (to the right)
            # If yes, we need to use the cell below as the actual day cell
            has_time_in_same_row = False
            for col in range(3, 15):  # Check columns C to N
                if is_time_cell(get_cell_value(row, col)):
                    has_time_in_same_row = True
                    break

            if has_time_in_same_row:
                # Check if the cell below also has the same day name
                # This handles the Monday case where B5 and B6 both have "Monday"
                cell_below = get_cell_value(row + 1, 2)
                if is_day_name(cell_below) and str(cell_below).strip().lower() == str(cell_value).strip().lower():
                    # Use the cell below as the day cell
                    day_row = row + 1
                    time_row = row
                    room_search_start = row + 2  # Rooms start after the duplicate day cell
                    processed_rows.add(row)
                    processed_rows.add(row + 1)
                else:
                    # Just use current row as day, time is in the same row
                    day_row = row
                    time_row = row
                    room_search_start = row + 1
                    processed_rows.add(row)
            else:
                # Look for "Rooms" cell below the day cell
                rooms_row = None
                for r in range(row + 1, min(row + 5, max_row + 1)):
                    rooms_value = get_cell_value(r, 2)
                    if rooms_value and str(rooms_value).strip().lower() == 'rooms':
                        rooms_row = r
                        break

                if rooms_row:
                    # Time row is the one with "Rooms"
                    time_row = rooms_row
                    room_search_start = rooms_row + 1
                else:
                    # Default: time row is one above the day cell
                    time_row = row - 1
                    room_search_start = row + 1

                processed_rows.add(row)

            # Detect time columns (columns that have time values in the time_row)
            time_cols = []
            for col in range(3, 20):  # Check columns C onwards
                cell_val = get_cell_value(time_row, col)
                if is_time_cell(cell_val):
                    time_cols.append(col)

            if not time_cols:
                continue  # Skip if no time columns found

            # Detect room rows (rows below the time_row that have room names in column B)
            room_rows = []
            # Use room_search_start if defined, otherwise default to time_row + 1
            for r in range(room_search_start, min(room_search_start + 20, max_row + 1)):
                room_value = get_cell_value(r, 2)

                # Stop if we hit another day name or empty rows
                if is_day_name(room_value):
                    break

                # Check if this row has room data (non-empty cell in column B)
                if room_value and str(room_value).strip():
                    # Skip if it's "Rooms" label
                    if str(room_value).strip().lower() != 'rooms':
                        room_rows.append(r)
                elif len(room_rows) > 0:
                    # Stop if we've found rooms and now hit an empty cell
                    # But allow a few empty cells in between
                    empty_count = 0
                    for check_r in range(r, min(r + 3, max_row + 1)):
                        if not get_cell_value(check_r, 2):
                            empty_count += 1
                    if empty_count >= 2:
                        break

            if room_rows:
                day_name = get_cell_value(day_row, 2)
                if isinstance(day_name, datetime):
                    day_name = day_name.strftime("%A")
                else:
                    day_name = str(day_name).strip()

                blocks.append({
                    "day_cell": f"B{day_row}",
                    "day_row": day_row,
                    "time_row": time_row,
                    "room_rows": room_rows,
                    "time_cols": time_cols,
                    "day_name": day_name  # For debugging
                })
                print(f"    - {day_name}: day_row={day_row}, time_row={time_row}, rooms={len(room_rows)}, time_cols={len(time_cols)}")

    return blocks

def convert_xlsx_to_csv(input_filename, output_filename=None):
    """
    Convert Excel timetable to CSV format with automatic detection of day blocks.

    Parameters:
    input_filename (str): Path to the input XLSX file
    output_filename (str, optional): Path to the output CSV file. If None,
                                    will use input filename with .csv extension

    Returns:
    str: Path to the created CSV file
    """
    # If output_filename not provided, derive from input_filename
    if output_filename is None:
        output_filename = os.path.splitext(input_filename)[0] + '.csv'

    # Ensure output directory exists
    output_dir = os.path.dirname(output_filename)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    # Load workbook and get all worksheets
    wb = load_workbook(filename=input_filename)
    worksheets = wb.worksheets

    # Automatically detect day blocks from all worksheets
    all_day_blocks = []
    for sheet_index, ws in enumerate(worksheets):
        print(f"Detecting blocks in sheet {sheet_index + 1}: {ws.title}")
        blocks = detect_day_blocks(ws)
        for block in blocks:
            block["sheet_index"] = sheet_index
            all_day_blocks.append(block)
        print(f"  Found {len(blocks)} day block(s)")

    print(f"\nTotal blocks detected across all sheets: {len(all_day_blocks)}")

    if not all_day_blocks:
        print("Warning: No day blocks detected in the workbook!")
        # Create empty CSV
        df = pd.DataFrame(columns=["Day", "Time", "Room", "Subject", "Class/Group", "Teacher(s) Name"])
        df.to_csv(output_filename, index=False)
        return output_filename

    def fix_class_group_format(class_group_str):
        if not class_group_str:
            return class_group_str
            
        # Format 1: BSDS/BSAI-6A → BSDS-6A & BSAI-6A
        class_group_str = re.sub(
            r'\b([A-Z]{4})/([A-Z]{4})-(\d+[A-Za-z]*)\b',
            r'\1-\3 & \2-\3',
            class_group_str
        )
        
        # Format 2: BSAI-1A/BSDS-1A → BSAI-1A & BSDS-1A
        class_group_str = re.sub(
            r'\b([A-Z]{4}-\d+[A-Za-z]*)/([A-Z]{4}-\d+[A-Za-z]*)\b',
            r'\1 & \2',
            class_group_str
        )

        # Format 3: BSSE-2A,2B,2C,2D,2E → BSSE-2A & BSSE-2B & BSSE-2C & BSSE-2D & BSSE-2E
        class_group_str = re.sub(
            r'([A-Z]{4}-\d+[A-Za-z]*)[,]([A-Z]{4}-\d+[A-Za-z]*)',
            r'\1 & \2',
            class_group_str
        )

        return class_group_str


    def get_cell_value(row, col):
        cell = ws.cell(row=row, column=col)
        coord = f"{get_column_letter(col)}{row}"
        for merged in merged_cells:
            if coord in merged:
                return ws.cell(merged.min_row, merged.min_col).value
        return cell.value

    # Extract structured timetable
    rows = []

    for block in all_day_blocks:
        # Get the appropriate worksheet
        ws = worksheets[block["sheet_index"]]

        # Get merged cells for this worksheet
        merged_cells = ws.merged_cells.ranges

        # Update get_cell_value to use current worksheet
        def get_cell_value(row, col):
            cell = ws.cell(row=row, column=col)
            coord = f"{get_column_letter(col)}{row}"
            for merged in merged_cells:
                if coord in merged:
                    return ws.cell(merged.min_row, merged.min_col).value
            return cell.value

        raw_day = get_cell_value(block["day_row"], 2)  # Column B = 2
        if isinstance(raw_day, datetime):
            day = raw_day.strftime("%A")
        else:
            day = str(raw_day).strip() if raw_day else "Unknown"

        time_row = block["time_row"]
        room_rows = block["room_rows"]
        time_cols = block["time_cols"]

        # Get all time slots for this block
        def clean_time_value(val):
            """Clean time value by removing tabs and normalizing dashes"""
            if not val:
                return ""
            time_str = str(val).strip()
            # Replace tabs with spaces
            time_str = time_str.replace('\t', ' ')
            # Replace various dash types with standard hyphen
            time_str = time_str.replace('–', '-').replace('—', '-')
            # Normalize spaces around dash
            time_str = re.sub(r'\s*-\s*', ' - ', time_str)
            return time_str

        time_slots = {col: clean_time_value(get_cell_value(time_row, col)) for col in time_cols}

        def get_merged_time_range(row, col):
            """Get the time range for a cell, considering if it's merged across multiple time columns"""
            coord = f"{get_column_letter(col)}{row}"

            # Check if this cell is part of a merged range
            for merged in merged_cells:
                if coord in merged:
                    # Get the start and end columns of the merged range
                    start_col = merged.min_col
                    end_col = merged.max_col

                    # Find the time columns that overlap with this merged range
                    overlapping_time_cols = [tc for tc in time_cols if start_col <= tc <= end_col]

                    if len(overlapping_time_cols) >= 2:
                        # Get start time from first column and end time from last column
                        first_time_col = overlapping_time_cols[0]
                        last_time_col = overlapping_time_cols[-1]

                        first_time_str = time_slots.get(first_time_col, "")
                        last_time_str = time_slots.get(last_time_col, "")

                        # Extract start time from first slot and end time from last slot
                        first_match = re.search(r'(\d{1,2}:\d{2})', first_time_str)
                        last_match = re.search(r'(\d{1,2}:\d{2})\s*$', last_time_str)

                        if first_match and last_match:
                            start_time = first_match.group(1)
                            end_time = last_match.group(1)
                            return f"{start_time} - {end_time}"
                    elif len(overlapping_time_cols) == 1:
                        # Only one time column, use its time
                        return time_slots.get(overlapping_time_cols[0], "")

            # Not merged, return the time for this column
            return time_slots.get(col, "")

        # Track which cells we've already processed (to avoid duplicates from merged cells)
        processed_cells = set()

        # Loop through each row (i.e., room) and each time slot
        for r in room_rows:
            room = get_cell_value(r, 2)  # Column B = 2 (Rooms)
            for c in time_cols:
                coord = f"{get_column_letter(c)}{r}"

                # Skip if we've already processed this cell (part of a merged range)
                if coord in processed_cells:
                    continue

                val = get_cell_value(r, c)
                if not val:
                    continue

                # Mark this cell and any merged cells as processed
                for merged in merged_cells:
                    if coord in merged:
                        # Mark all cells in this merged range as processed
                        for mr in range(merged.min_row, merged.max_row + 1):
                            for mc in range(merged.min_col, merged.max_col + 1):
                                if mc in time_cols:  # Only mark time columns
                                    processed_cells.add(f"{get_column_letter(mc)}{mr}")
                        break
                else:
                    # Not merged, just mark this cell
                    processed_cells.add(coord)

                # Handle multi-line and complex cell content
                lines = [line.strip() for line in str(val).strip().split("\n") if line.strip()]

                # Skip unwanted entries
                if lines and lines[0].lower() in ["used in cs department", "namaz break"]:
                    continue

                subject = lines[0]
                groups = []
                teachers = []
                teacher_timeslot = None

                for line in lines[1:]:
                    # Detect group identifiers
                    if re.search(r'\b(BS\w{2,4})[-/]\d+[A-Za-z]*', line):
                        # Convert "BSDS/BSAI-6A" to "BSDS-6A & BSAI-6A"
                        line = fix_class_group_format(line)
                        groups.append(line)
                    # Check if this line is just a time slot in parentheses (e.g., "(10:45 am to 12:25 pm)")
                    elif re.match(r'^\s*\(([^)]+)\)\s*$', line):
                        timeslot_text = re.match(r'^\s*\(([^)]+)\)\s*$', line).group(1).strip()

                        # Extract time from various formats and convert to 24-hour
                        time_match = re.search(r'(\d{1,2}:\d{2})\s*(AM|PM)?\s*(?:-|TO|to)\s*(\d{1,2}:\d{2})\s*(AM|PM)?', timeslot_text, re.IGNORECASE)
                        if time_match:
                            start_time = time_match.group(1)
                            start_period = time_match.group(2)
                            end_time = time_match.group(3)
                            end_period = time_match.group(4)

                            # Convert to 24-hour format if AM/PM is present
                            if start_period:
                                start_time = convert_to_24hour(start_time, start_period)
                            if end_period:
                                end_time = convert_to_24hour(end_time, end_period)

                            teacher_timeslot = f"{start_time} - {end_time}"
                    else:
                        # Detect teacher and timeslot on same line (e.g., "Ms. Namra Amjad (10:45 AM TO 12:25 PM)")
                        match = re.match(r"(.*?)\s*\(([^)]+)\)\s*$", line)
                        if match:
                            teacher_name = match.group(1).strip()
                            timeslot_text = match.group(2).strip()

                            # Extract time from various formats and convert to 24-hour
                            time_match = re.search(r'(\d{1,2}:\d{2})\s*(AM|PM)?\s*(?:-|TO|to)\s*(\d{1,2}:\d{2})\s*(AM|PM)?', timeslot_text, re.IGNORECASE)
                            if time_match:
                                start_time = time_match.group(1)
                                start_period = time_match.group(2)
                                end_time = time_match.group(3)
                                end_period = time_match.group(4)

                                # Convert to 24-hour format if AM/PM is present
                                if start_period:
                                    start_time = convert_to_24hour(start_time, start_period)
                                if end_period:
                                    end_time = convert_to_24hour(end_time, end_period)

                                teacher_timeslot = f"{start_time} - {end_time}"

                            teachers.append(teacher_name)
                        else:
                            # Consider it part of teacher names if not matched with timeslot
                            teachers.append(line)

                # If a timeslot was detected in the teacher's name, use it
                # Otherwise, get the time range considering merged cells
                if teacher_timeslot:
                    time_for_this_class = teacher_timeslot
                else:
                    time_for_this_class = get_merged_time_range(r, c)

                # Apply fix_class_group_format to the entire group string
                full_groups = " & ".join(groups)  # Now this will have the proper format
                full_teachers = " ".join([t.strip().strip(",") for t in ", ".join(teachers).split(",") if t.strip()])

                # Append row with formatted data
                rows.append({
                    "Day": day,
                    "Time": time_for_this_class,
                    "Room": str(room).strip(),
                    "Subject": subject,
                    "Class/Group": fix_class_group_format(full_groups),  # Apply here
                    "Teacher(s) Name": full_teachers
                })

    # Convert to DataFrame
    df = pd.DataFrame(rows)

    def fix_time_format(time_str):
        if not time_str:
            return time_str
        # Replace tabs with a dash if they appear between times
        # Pattern: "HH:MM<tab>HH:MM" -> "HH:MM - HH:MM"
        time_str = re.sub(r'(\d{1,2}:\d{2})\s*\t\s*(\d{1,2}:\d{2})', r'\1 - \2', time_str)
        # Replace remaining tabs with spaces
        time_str = time_str.replace('\t', ' ')
        # Replace various dash types (en-dash, em-dash, hyphen) with standard hyphen
        time_str = time_str.replace('–', '-').replace('—', '-')
        # Normalize all time ranges to have exactly one space before and after the dash
        return re.sub(r'\s*-\s*', ' - ', time_str.strip())

    # Apply fixes
    df['Time'] = df['Time'].apply(fix_time_format)
    df['Class/Group'] = df['Class/Group'].apply(fix_class_group_format)

    # Output
    df.to_csv(output_filename, index=False)
    print(f"Converted {input_filename} to {output_filename}")
    
    return output_filename

def convert_to_24hour(time_str, period):
    """Convert 12-hour time to 24-hour format"""
    if not period:
        return time_str
    
    hour, minute = time_str.split(':')
    hour = int(hour)
    
    if period.upper() == 'PM' and hour != 12:
        hour += 12
    elif period.upper() == 'AM' and hour == 12:
        hour = 0
    
    return f"{hour:02d}:{minute}"

# If run directly, use the hardcoded filename
if __name__ == "__main__":
    input_file = 'Timetable SE Department (Fall-25) UpdatedVersion - 1.1.xlsx'
    convert_xlsx_to_csv(input_file)
